from time import strftime
from typing import Iterable

from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

from src import NONBLANK, VALUESET
from strtools import str_capture_draw, str_union, strf_time_ts

"""
Plugin interface spec

For data collection functions:
Input: Worksheet, tuple[Cell], <arguments[dict]>
Output: Generator which yields tuple[Iterable], each Iterable stands for a column

For data manipulation functions:
Input: source tuple[Iterable]
Output: Generator which yields tuple[Iterable], each Iterable stands for a row
"""

STR_CONCAT_MODE = "字符串合并连接模式"
STR_FORMAT_MODE = "字符串按列头格式化模式"
STR_FORMAT_CAPS = "字符串按格式化列头"
CAP_DRW_MODES = "匹配抽取模式"

STAROW, WRAPROW, MAXROW = "起始行", "包裹行", "最大行"
IN_FORMAT, DT_FORMAT = "输入格式", "日期格式"
DC_COL_NS, DC_ID_KV = "收支列名", "收支标值{目标键:值,...}"




def determine_row_bands(args: dict, start: int, last: int) -> tuple:
    sta_row = start
    max_row = args[MAXROW] if args and MAXROW in args else last
    end_row = max_row - args[WRAPROW] if args and WRAPROW in args and args[WRAPROW] else max_row

    return sta_row, max_row, end_row






def ConstantInject(ws: Worksheet, cells: tuple[Cell]=None, meta_args: dict=None, field_args: dict=None) -> Iterable:

    assert meta_args and STAROW in meta_args, f"常量注入插件应指定 {STAROW} 参数。"
    assert field_args and VALUESET in field_args, f"常量注入插件应指定 {VALUESET} 参数。"

    sta_row, max_row, end_row = determine_row_bands(meta_args, meta_args[STAROW], ws.max_row)
    print(f"[ConstantInject]:: 为 {sta_row} 至 {end_row} 行注入匹配数据（[匹配抽取]可选）。")

    cons_v = field_args[VALUESET]

    for i in range(end_row - sta_row):
        yield tuple(cons_v)



def OneShot(ws: Worksheet, cells: tuple[Cell]=None, meta_args: dict=None, field_args: dict=None) -> Iterable:
    if field_args and CAP_DRW_MODES in field_args:
        rgexp: str = field_args[CAP_DRW_MODES]
        assert rgexp, f"{CAP_DRW_MODES} 参数必须是非空的有效值。"
    else:
        rgexp: None = None

    if field_args and STR_CONCAT_MODE in field_args:
        concat_mod: str = field_args[STR_CONCAT_MODE]
    else:
        concat_mod: None = None

    col_caps = ()
    if field_args and STR_FORMAT_MODE in field_args:
        format_mod: str = field_args[STR_FORMAT_MODE]
        assert format_mod, f"{STR_FORMAT_MODE} 参数必须是非空的有效值。"

        if STR_FORMAT_CAPS in field_args:
            col_caps: tuple = tuple(field_args[STR_FORMAT_CAPS])
        else:
            assert False, "[OneShot] 插件格式化字符串，必须指定列头列表参数。"

        assert isinstance(col_caps, tuple) and len(col_caps) > 0, f"{STR_FORMAT_CAPS} 参数的有效值必须是非空的列表。"
    else:
        format_mod: None = None

    if field_args and NONBLANK in field_args:
        non_blank: bool = field_args[NONBLANK]
        assert rgexp, f"{NONBLANK} 参数必须是非空的有效值。"
    else:
        non_blank: bool = False

    assert meta_args and STAROW in meta_args, f"一次性读取插件应指定 {STAROW} 参数(必填)。"

    sta_row, max_row, end_row = determine_row_bands(meta_args, meta_args[STAROW], ws.max_row)
    print(f"[OneShot]:: 从 {cells[0].coordinate + (':' + cells[-1].coordinate if len(cells) > 1 else '')} "
          f"单元格为 {sta_row} 至 {end_row} 行生成匹配数据。")

    col_vals = tuple(c.value for c in cells)

    if non_blank and not any(col_vals):
        return # !! This will generates an empty tuple

    if concat_mod or format_mod:
        col_vals = str_union(col_vals, col_caps, concat_mod, format_mod)

    if rgexp:
        col_vals = str_capture_draw(col_vals, rgexp)

    for i in range(end_row - sta_row):
        yield col_vals


def ColumnHead(ws: Worksheet, cells: tuple[Cell]=None, meta_args: dict=None, field_args: dict=None) -> Iterable:
    if field_args and CAP_DRW_MODES in field_args:
        rgexp: str = field_args[CAP_DRW_MODES]
        assert rgexp, f"{CAP_DRW_MODES} 参数必须是非空的有效值。"
    else:
        rgexp: None = None

    if field_args and NONBLANK in field_args:
        non_blank: bool = field_args[NONBLANK]
    else:
        non_blank: bool = False

    com_row = 0
    for row in (c.row for c in cells):
        if com_row == 0:
            com_row = row
        assert row == com_row, f"指定的列头应当位于同一列中 {com_row}。"

    cols = tuple(c.column for c in cells)
    min_col = min(cols)
    max_col = max(cols)

    sta_row, max_row, end_row = determine_row_bands(meta_args, com_row + 1, ws.max_row)
    print(f"[ColumnHead]:: 从 {sta_row} 行开始加载 {str(cols)} 等列数据。")

    for values in ws.iter_rows(sta_row, end_row, min_col, max_col, values_only=True):
        col_vals = tuple(values[col - min_col] for col in cols)
        try:
            if rgexp:
                col_vals = str_capture_draw(col_vals, rgexp)
        except ValueError:
            assert not non_blank, f"'非空列取得空值，数据：{''.join(col_vals)}':捕获'{rgexp}' "

        yield col_vals


def ColumnUnited(ws: Worksheet, cells: tuple[Cell]=None, meta_args: dict=None, field_args: dict=None) -> Iterable:

    if field_args and CAP_DRW_MODES in field_args:
        rgexp: str = field_args[CAP_DRW_MODES]
        assert rgexp, f"{CAP_DRW_MODES} 参数必须是非空的有效值。"
        inhr_args = field_args.copy()
        inhr_args.pop(CAP_DRW_MODES)
    else:
        inhr_args = field_args
        rgexp: None = None

    if field_args and STR_CONCAT_MODE in field_args:
        concat_mod: str = field_args[STR_CONCAT_MODE]
    else:
        concat_mod: None = None

    if field_args and STR_FORMAT_MODE in field_args:
        format_mod: str = field_args[STR_FORMAT_MODE]
        assert format_mod, f"{STR_FORMAT_MODE} 参数必须是非空的有效值。"
    else:
        format_mod: None = None

    col_caps = [c.value for c in cells]
    for col_vals in ColumnHead(ws, cells, meta_args, inhr_args):
        ts = tuple(str_union(col_vals, col_caps, concat_mod, format_mod))
        if rgexp:
            yield str_capture_draw(ts, rgexp)
        else:
            yield ts



def IdentifyByColumn(ws: Worksheet, cells: tuple[Cell]=None, meta_args: dict=None, field_args: dict=None) -> Iterable:

    dc_map = None
    if field_args and DC_ID_KV in field_args:
        dc_map = field_args[DC_ID_KV]
    assert dc_map, f"CreditDebitByColumn插件, 参数:{DC_ID_KV} 必须配置。"

    non_blank = field_args[NONBLANK] if field_args and NONBLANK in field_args else False

    cell_names = [c.value for c in cells]
    try:
        cell_indxes = tuple(cell_names.index(k) for k in dc_map.keys())

        for row in ColumnHead(ws, cells, meta_args=meta_args, field_args=field_args):
            for i in cell_indxes:
                if not row[i]:
                    continue
                yield (dc_map[cell_names[i]], )
                break
            else:
                assert not non_blank, f"'非空列取得空值，数列：{','.join(str(cell_indxes))}。"
                yield ('', )
    except ValueError as err:
        assert False, f"{DC_ID_KV} 配置：'{str(err)}'与配置列头 {':'.join(c.coordinate for c in cells)} 的取值不符。"



def IdentifyByValue(ws: Worksheet, cells: tuple[Cell]=None, meta_args: dict=None, field_args: dict=None) -> Iterable:

    assert field_args and DC_ID_KV in field_args, f"使用'CreditDebitByValue'必须指定 {DC_ID_KV} 收支映射项。"
    dc_map = field_args[DC_ID_KV]

    non_blank = field_args[NONBLANK] if field_args and NONBLANK in field_args else False

    for row in ColumnHead(ws, cells, meta_args=meta_args, field_args=field_args):
        k = ' '.join(row)
        try:
            yield (dc_map[k],)
        except KeyError as e:
            if non_blank:
                break
            yield ('', )


def DateTimeFormator(ws: Worksheet, cells: tuple[Cell]=None, meta_args: dict=None, field_args: dict=None) -> Iterable:

    assert field_args and IN_FORMAT in field_args and field_args[IN_FORMAT], f"{IN_FORMAT} 参数必须指定且不应该为空。"
    dt_format = field_args[IN_FORMAT]

    assert field_args and DT_FORMAT in field_args and field_args[DT_FORMAT], f"{DT_FORMAT} 参数必须指定且不应该为空。"
    st_format = field_args[DT_FORMAT]

    non_blank = field_args[NONBLANK] if field_args and NONBLANK in field_args else False

    for row in ColumnHead(ws, cells, meta_args=meta_args, field_args=field_args):
        try:
            yield (strftime(st_format, strf_time_ts(row, format = dt_format)),)
        except ValueError:
            if non_blank:
                break
            yield ('', )


def MergedDateTime(sources: tuple[Iterable], arguments: dict=None) -> Iterable:

    if arguments and IN_FORMAT in arguments:
        dt_format = arguments[IN_FORMAT]
        assert dt_format, f" {IN_FORMAT} 参数必须指定且不应该为空。"

    for l, r in sources:
        assert l, f"合并日期时间插件要求日期不能为空，{l} {r}"
        yield (strftime("%Y-%m-%d %H:%M:%S", strf_time_ts(l, r)),)


