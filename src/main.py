from csv import reader as csv_reader
from io import BufferedReader, BufferedIOBase
from os import getcwd, listdir
from os.path import join as joinpath, splitext, basename, isdir, isfile
from re import match
from traceback import print_exc
from zipfile import is_zipfile as is_zip, Path, ZipFile

from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from tomli import TOMLDecodeError
from xls2xlsx import XLS2XLSX

from actuator import run_policy
from loader import load, get_meta, get_policies, get_out_uri, find_matching_policy
from plugin import load_plugin
from reader import read_xlsx
from src import METADATA, POLICYNAME, FINISH, DATASOURCE, ARGUMENT, PLUGIN, VALUESET
from writer import write_file


def csv_converter(file_name: str, fd: BufferedReader|BufferedIOBase):
    wb = Workbook()
    ws = wb.create_sheet(splitext(file_name)[0])

    for row in csv_reader(fd):
        line = [cell.replace('\t', '') if isinstance(cell, str) else cell for cell in row]
        ws.append(line)

    return wb

def read_table_file(name: str, meta: dict, confs: dict, file_name: str, fd: BufferedReader|BufferedIOBase) -> dict:

    if file_name.endswith('.xlsx'):
        wb = load_workbook(fd)
        assert wb
    elif file_name.endswith('.csv'):
        wb = csv_converter(file_name, fd)
        assert wb
    elif file_name.endswith('.xls'):
        try:
            wb = XLS2XLSX(fd).to_xlsx()
        except ValueError as e:
            if len(e.args) > 0 and e.args[0].__class__.__name__ == 'XLRDError':
                wb = load_workbook(fd)
    else:
        raise Exception(f"{file_name} 既不是 .xlsx/.csv 文件，也不是旧版本 .xls 文件，不能读取。")

    sources = read_xlsx(name, wb, meta, confs)
    return sources


def post_masquerade(meta: dict, sources: dict) -> dict:
    if not (FINISH in meta and meta[FINISH]):
        return sources

    post_procs = meta[FINISH]

    for post_proc in post_procs:
        assert PLUGIN in post_proc and post_proc[PLUGIN], f"后处理过程要求必须指定插件."
        plugin_name = post_proc[PLUGIN]

        assert DATASOURCE in post_proc and post_proc[DATASOURCE], f"后处理过程要求必须指定数据源."
        ds = post_proc[DATASOURCE]

        assert VALUESET in post_proc and post_proc[VALUESET], f"后处理过程要求必须指定 {VALUESET}。"
        vs = post_proc[VALUESET]

        fn = load_plugin(plugin_name)
        arguments = None
        if ARGUMENT in post_proc:
            arguments = post_proc[ARGUMENT]
            assert arguments, f"完成组装 {plugin_name} 插件不能使用参数 '{str(arguments)}'"
        values = fn((zip(ds, ls) for ls in zip(*(sources.pop(k) for k in ds))), arguments)

        sources.update(dict(zip(vs, zip(*(v for v in values)))))

    return sources


def prune_extra_length(data_set: dict) -> tuple:
    min_len, max_len = None, 0
    try:
        for d_key, d_lst in data_set.items():
            d_lst = tuple(d_lst)
            d_len = len(d_lst)

            if min_len is None or d_len < min_len:
                min_len = d_len

            if d_len > max_len:
                max_len = d_len

            data_set[d_key] = d_lst
            print(f"'{d_key}'列载入数据 {min_len} 行 {f", 丢弃数据 {d_len - min_len} 行" if d_len > min_len else ""}。")
        else:
            if min_len is None:
                min_len = 0

        for d_key, d_lst in data_set.items():
            data_set[d_key] = d_lst[:min_len]

        return min_len, max_len - min_len
    except ValueError:
        assert False, "获取到的数据异常，这通常表示与位置有关的配置项错误。"



def assemble_data_pack(result: list[dict], data: dict, value_set: list[str]|tuple[str]):

    data_set = dict()
    for key in value_set:
        try:
            data_set[key] = data.pop(key)
        except KeyError:
            assert False, f"生成的数据集字段错误, 缺少 {key} 字段。"
    else:
        ac = len(data.keys())
        if ac > 0:
            print(f"丢弃 {ac} 个字段： {str(tuple(data.values()))}")

    result.append(data_set)

def is_xl_file(file_name: str) -> bool:
    return file_name.endswith('.xls') or file_name.endswith('.xlsx')

def is_csv_file(file_name: str) -> bool:
    return file_name.endswith('.csv')

def is_zip_file(file_path: str) -> bool:
    return is_zip(file_path)

def xls_in_zip(ar_path:str):
    with ZipFile(ar_path) as ar:
        for member_name in ar.namelist():
            arp = Path(ar, member_name)
            if arp.is_dir():
                continue

            file_name = basename(member_name)
            if file_name.startswith(r'.') or (file_name.startswith(r'__') and file_name.endswith(r'__')):
                continue

            with ar.open(member_name, 'r') as fd:
                yield fd, f"{basename(ar_path)}/{member_name}"


def xls_in_dir(xl_dir: str):

    for file_name in listdir(xl_dir):
        if file_name.startswith(r'.') or (file_name.startswith(r'__') and file_name.endswith(r'__')):
            continue

        file_path = joinpath(xl_dir, file_name)
        if isfile(file_path):
            if is_xl_file(file_name):
                with open(file_path, 'rb') as fd:
                    yield fd, file_name
            elif is_csv_file(file_path):
                with open(file_path, 'rt') as fd:
                    yield fd, file_name
            elif is_zip_file(file_path):
                yield from xls_in_zip(file_path)

        elif isdir(file_path):
            yield from xls_in_dir(file_path)

        else:
            print(f" {file_name} 不是 xl、zip文件或文件夹，已忽略。")
            continue


def main(source: str='xls', out: str='xls', conf: str=''):
    inc_files = list()
    exc_files = list()
    success = False

    try:
        delivered_data_packs: list[dict] = list()
        root_dir = getcwd()

        try:
            acc_conf = load("/".join((root_dir, conf)))
            conf_meta = get_meta(acc_conf)
            out_path = get_out_uri(conf_meta, f"{root_dir}\\{out}")
        except TOMLDecodeError as e:
            assert False, f"配置文件发生错误 - {e.args}"

        assert VALUESET in conf_meta and conf_meta[VALUESET], f" 必须指定根 {VALUESET}。"
        v_set = conf_meta[VALUESET]

        policies = get_policies(acc_conf)
        print("配置文件加载完成。")

        for xl_filed, xl_name in xls_in_dir(f"{root_dir}\\{source}"):
            print(f"  ==: 开始处理工作簿：{xl_name} :==  ")

            for (abb_name, acc_des) in acc_conf.items():
                if not match(abb_name, xl_name):
                    continue

                acc_meta = get_meta(acc_des)
                bundle = read_table_file(abb_name, acc_meta, acc_des, xl_name, xl_filed)

                assert POLICYNAME in acc_meta, f"{abb_name}账户配置的{METADATA}中应包含适用的策略。"
                policy_name = acc_meta[POLICYNAME]

                policy = find_matching_policy(policy_name, policies)
                if not policy:
                    print(f"没有策略匹配 '{abb_name}'.'{METADATA}' 所指定的 '{policy_name}' 项。")
                    break

                pol_meta = get_meta(policy)
                collection = run_policy(policy, bundle, policy_name, abb_name)

                record_cnt, prune_cnt = prune_extra_length(collection)
                if record_cnt > 0:
                    pack = post_masquerade(pol_meta, collection)
                    assemble_data_pack(delivered_data_packs, pack, v_set)

                inc_files.append((xl_name, record_cnt, prune_cnt))
                print(f"  ==: {xl_name} 处理完成！ :==  \n")
                break
            else:
                exc_files.append(xl_name)
                print(f"  >>: 没有为 {xl_name} 配置描述项，已忽略！ :<<  ")

        write_file(delivered_data_packs, out_path, v_set)
        success = True
    except AssertionError as asr_err:
        print(f"《《《错误中止》》》： {str(asr_err)}")
    except BaseException as bas_exp:
        print(str(bas_exp))
        print_exc()

    print(f"\n == 总结 == \n任务运行'{"完毕" if success else "中断"}', 共处理 {len(inc_files)} 个文件。"
          f"{f"，发现未处理{len(exc_files)}个：{str(exc_files)}; " if len(exc_files) else ""}。\n数据未完全归集文件情况：")
    for fr in inc_files:
        if fr[-1] != 0:
            print(f"\t - {fr[0]} 获取 {fr[1]} 行，丢弃{fr[-1]} 行。")
    input("回车退出。")