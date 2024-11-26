from re import match

from openpyxl.cell import Cell, MergedCell
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from loader import attributes
from plugin import load_plugin
from src import ADDRSET, PLUGIN, ARGUMENT, WSNAME_PATTERN


def read_xlsx(abb_name: str, wb: Workbook, meta:dict, acc_des: dict):
    wsn_ptn = ""
    if WSNAME_PATTERN in meta:
        wsn_ptn = meta[WSNAME_PATTERN]

    for ws_name in wb.sheetnames:
        if wsn_ptn:
            if match(wsn_ptn, ws_name):
                ws = wb[ws_name]
                return read_xl_sheet(ws, meta, acc_des)
        else:
            try:
                ws = wb[wb.sheetnames[0]]
            except BaseException as err:
                assert False, f"{abb_name} {WSNAME_PATTERN} 参数未配置，加载默认工作表错误，原因 {err.args}"
            return read_xl_sheet(ws, meta, acc_des)
    else:
        raise Exception(f"在所有表单中，都没有找到符合 {wsn_ptn} 命名的表单，检查配置是否错误。")


def read_xl_sheet(ws: Worksheet, meta: dict, attrs: dict) -> dict:

    binding = dict()
    for key, attr in attributes(attrs):

        assert ADDRSET in attr, f"字段 '{key}' 缺失 '{ADDRSET}' 属性配置项。"
        addr = attr[ADDRSET]

        assert PLUGIN in attr, f"字段 '{key}' 缺失 '{PLUGIN}' 属性配置项。"
        plugin_name = attr[PLUGIN]

        binding[key] = bind_with_cells(ws, plugin_name, addr, meta, attrs[key])

    return binding


def bind_with_cells(ws: Worksheet, plugin: str, addrs: tuple[str], meta: dict, attr: dict) -> dict:

    if ARGUMENT in meta:
        meta_args = meta[ARGUMENT]
    else:
        meta_args = None

    if ARGUMENT in attr:
        field_args = attr[ARGUMENT]
    else:
        field_args = None

    return apply_plugin_to_cells(ws, plugin, addrs, meta_args=meta_args, field_args=field_args)


def apply_plugin_to_cells(ws: Worksheet, plugin_name: str, addrs: tuple[str], meta_args: dict=None, field_args: dict=None):
    fn = load_plugin(plugin_name)

    if addrs:
        cells = flatten_cells(ws, addrs)

        return fn(ws, cells=cells, meta_args=meta_args, field_args=field_args)

    return fn(ws, meta_args=meta_args, field_args=field_args)


def flatten_cells(ws: Worksheet, addrs: tuple[str]) -> list[Cell]:
    cset = list()

    for addr in addrs:
        try:
            cells = ws[addr]
        except IndexError:
            assert False, f"配置错误 {addr} 不是一个有效的 xl 地址。"

        if isinstance(cells, Cell) or isinstance(cells, MergedCell):
            cset.append(cells)

        elif isinstance(cells, tuple):
            for r in cells:
                if isinstance(r, Cell):
                    cset.append(r)
                elif isinstance(r, tuple):
                    cset.extend(c for c in r)
                else:
                    assert False, f"加载地址 {str(addrs)} 时出现内部错误。"

    assert len(cset), f"{str(addrs)} 不是一个有效的地址集。"

    return cset